import sys
from selenium import webdriver
import time
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager




#ブックを取得

wb = openpyxl.load_workbook('matrix.xlsx')
#シートを取得

ws = wb["Sheet1"]
#import pandas as pd


USER = ws["A8"].value
PASS = ws["A9"].value


#GoogleChromeを起動


#options = webdriver.chrome.options.Options()
#options.binary_location = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'
#browser = webdriver.Chrome(options=options)
browser = webdriver.Chrome(ChromeDriverManager().install())
#browser = webdriver.Chrome(executable_path = './driver/chromedriver')
browser.implicitly_wait(3)

#ログインページにアクセスするサイトへアクセス
url_login = "https://portal.nap.gsic.titech.ac.jp/GetAccess/Login?Template=userpass_key&AUTHMETHOD=UserPassword"
browser.get(url_login)
time.sleep(0)
print("ログインページにアクセスしました")


#テキストボックス入力

element = browser.find_element_by_name('usr_name')
element.clear()
element.send_keys(USER)
element = browser.find_element_by_name('usr_password')
element.clear()
element.send_keys(PASS)
print("フォームを送信")


#ログインボタンをおす
browser_from = browser.find_element_by_name("OK")
time.sleep(0)
browser_from.click()
print("情報を入力してログインボタンを押しました")



text1 = browser.find_element_by_name("login").text

#マトリクスコード解析

# i=0
# for text in text1:
#   print(text,i)
#   i = i+1





def choice(x):
    a = text1[x]+text1[x+2]
    saisixyo=ws[a].value
    return saisixyo


# print(choice(30))
# print(choice(36))
# print(choice(42))

message3=choice(30)
message4=choice(36)
message5=choice(42)


element = browser.find_element_by_name('message3')
element.clear()
element.send_keys(message3)
element = browser.find_element_by_name('message4')
element.clear()
element.send_keys(message4)
element = browser.find_element_by_name('message5')
element.clear()
element.send_keys(message5)
print("フォームを送信")
#ログインボタンをおす
browser_from = browser.find_element_by_name("OK")
time.sleep(0)
browser_from.click()
print("情報を入力してログインボタンを押しました")
print("ログイン完了しました")
